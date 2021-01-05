from dataknead.baseloader import BaseLoader
from .util import get_fieldnames, sniff_type, SnifferType
import openpyxl

class ExcelLoader(BaseLoader):
    EXTENSIONS = ["xlsx"]

    # Convert a openpyxl sheet to a dict / list we can use to convert to
    # other formats like json/csv
    def _sheet_to_dict(sheet, has_header):
        data = []
        rows = list(sheet.rows)

        if has_header:
            # If has_header is True, we assume the first row is a header
            header = [c.value for c in rows[0]]

            for row in rows[1:]:
                values = [c.value for c in row]
                data.append( dict(zip(header, values)) )
        else:
            # Otherwise, just return a simple list of lists
            for row in rows:
                data.append([ c.value for c in row ])

        return data

    @staticmethod
    def read(f, has_header = False, **kwargs):
        wb = openpyxl.load_workbook(f.name, **kwargs)

        if len(wb.sheetnames) == 1:
            # Only a single sheet, so return that as a list
            return ExcelLoader._sheet_to_dict(wb.active, has_header)
        else:
            # Multiple sheets, return as a dict with sheet names as keys
            data = {}

            for sheet in wb:
                data[sheet.title] = ExcelLoader._sheet_to_dict(sheet, has_header)

            return data

    @staticmethod
    def write(f, data):
        # We use the same logic here for writing to an Excel file as with CSV's
        wb = openpyxl.Workbook()
        sheet = wb.active

        data_type = sniff_type(data)

        # We wrap the append method so that we can catch errors and save
        # a string representation if all else failes
        def append(row):
            try:
                sheet.append(row)
            except Exception as e:
                print(f"Could not save data properly to Excel, saving a string representation")
                sheet.append([ str(row) ])

        if data_type == SnifferType.list_with_dicts:
            append(get_fieldnames(data))

            for row in data:
                append(list(row.values()))
        elif data_type in (SnifferType.list_with_lists, SnifferType.single_list):
            for row in data:
                append(row)
        elif data_type == SnifferType.single_dict:
            append(list(data.keys()))
            append(list(data.values()))

        wb.save(f.name)